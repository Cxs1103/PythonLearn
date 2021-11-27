#!/usr/bin/env python
# encoding: utf-8
'''
@File  : concatExcel.py
@Date  : 2021/11/27 0:30
@Author: Cxs
@Email: Cxs1103@163.com
@Desc  : Python批量合并多格式相同Excel文件，到一个表中
'''
import os
import openpyxl

dir = './datas'
# 获取目录下所有的表
origin_file_list = os.listdir(dir)

wb2 = openpyxl.Workbook()
for file in origin_file_list:
    file_path = dir + '/' + file
    sheet_name = file.split('/')[-1].split('.')[0][0:2]
    print(sheet_name)
    old_wb = openpyxl.load_workbook(file_path)
    old_sheet_name = old_wb.get_sheet_names()[0]
    old_ws = old_wb[old_sheet_name]
    ws2 = wb2.create_sheet(sheet_name)
    for row in old_ws.values:
        ws2.append(row)

execel_name = 'result.xlsx'

wb2.save(execel_name)
wb2.close() # 对程序中只读的workbook
