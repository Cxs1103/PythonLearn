#!/usr/bin/env python
# encoding: utf-8
'''
@File  : concatExcel.py
@Date  : 2021/11/27 0:30
@Author: Cxs
@Email: Cxs1103@163.com
@Desc  : Python批量合并多格式相同Excel文件，到一个表中
'''
import pandas as pd
import os
import shutil
import openpyxl

# 搜索目录
dir_path = "D:\work\Project"

# 将搜索到文件保存到列表中
fname_list = []
for root, dirs, files in os.walk(dir_path):
    for fname in files:
        if "配置信息表.xlsx" in fname:
            file_path = root + '/' + fname
            fname_list.append(os.path.join(root, fname))

#
wb2 = openpyxl.Workbook()
for files_name in fname_list:

    # 复制一份文件，进行操作
    target = './datas'
    if os.path.exists(target) == False:
        os.mkdir(target)
    shutil.copy(files_name, target)

    # print(files_name.split('.')[-2][0:2])
    # 创建sheet名称
    sheet_name = files_name.split('.')[-2][0:2]
    # 导入excel表格
    old_wb = openpyxl.load_workbook(files_name)
    # 获取第一个sheet的名称
    old_sheet_name = old_wb.get_sheet_names()[0]
    # 加载第一个sheet的内容
    old_ws = old_wb[old_sheet_name]
    ws2 = wb2.create_sheet(sheet_name)
    for row in old_ws.values:
        # 写入数据到新的sheet中
        ws2.append(row)

# 新excel名称
execel_name = 'result.xlsx'
# 保存文件
wb2.save(execel_name)
# 关闭文件
wb2.close()